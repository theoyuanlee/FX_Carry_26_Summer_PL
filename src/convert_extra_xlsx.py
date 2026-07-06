"""Convert the hand-pulled Bloomberg Excel (data/raw/FX_extra_data.xlsx) into the
same wide/long parquet format as the automated `bloomberg_data.py` groups.

Why this exists:
The Phase-2 supplemental groups could not be downloaded on this machine (no local
`blpapi`), so they were pulled by hand in Excel and saved as FX_extra_data.xlsx.
This script turns that workbook into `data/raw/{group}_{wide,long}.parquet` files
that are indistinguishable from the Bloomberg-downloaded groups, so
`fx_utils.load_wide` and the notebook treat them identically.

Workbook layout (one sheet per group):
- Column A lists the tickers top-to-bottom (labels start at row 2).
- Ticker k (0-indexed) has its (date, value) block at 1-indexed columns
  (3 + 3k, 4 + 3k) -> C/D, F/G, I/J, ...; a blank spacer column sits between
  blocks. Data starts at row 1 (the A-column labels are offset one row down and
  are only used for ordering: A-order == block-order).
- Blocks have independent calendars, so each is aligned on its own date column
  and the group is assembled by an outer join on the union of dates.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

SCRIPT_DIR = Path(__file__).resolve().parent
RAW_DIR = SCRIPT_DIR.parent / "data" / "raw"
XLSX_PATH = RAW_DIR / "FX_extra_data.xlsx"

FIELD = "PX_LAST"


def _read_sheet(ws) -> pd.DataFrame:
    """Return a wide DataFrame (index=date, columns=MultiIndex[ticker, field])."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    max_col = max(len(r) for r in rows)

    # Tickers in column A, from row 2 down (labels are offset one row below the data).
    tickers = [r[0] for r in rows[1:] if len(r) >= 1 and r[0] not in (None, "")]

    n_blocks = (max_col - 1) // 3  # blocks live at columns 3+3k / 4+3k (1-indexed)
    if len(tickers) != n_blocks:
        raise ValueError(
            f"sheet '{ws.title}': {len(tickers)} tickers in col A but {n_blocks} "
            f"data blocks (max_col={max_col}). Layout assumption broken."
        )

    series: dict[tuple[str, str], pd.Series] = {}
    for k, ticker in enumerate(tickers):
        d_idx = (3 + 3 * k) - 1  # 0-indexed date column
        v_idx = d_idx + 1        # 0-indexed value column
        dates, vals = [], []
        for r in rows:
            if len(r) <= v_idx:
                continue
            d, v = r[d_idx], r[v_idx]
            if d is None or v is None:
                continue
            dates.append(d)
            vals.append(v)
        s = pd.Series(
            pd.to_numeric(vals, errors="coerce"),
            index=pd.to_datetime(dates),
            name=(ticker, FIELD),
        )
        s = s[~s.index.duplicated(keep="last")].sort_index()
        series[(ticker, FIELD)] = s

    wide = pd.DataFrame(series)
    wide.columns = pd.MultiIndex.from_tuples(wide.columns, names=["ticker", "field"])
    wide.index.name = "date"
    return wide.sort_index()


def convert() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Expected workbook at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"Converting {XLSX_PATH.name}: sheets = {wb.sheetnames}")

    for group in wb.sheetnames:
        wide = _read_sheet(wb[group])
        if wide.empty:
            print(f"  {group}: EMPTY, skipped")
            continue

        wide_path = RAW_DIR / f"{group}_wide.parquet"
        long_path = RAW_DIR / f"{group}_long.parquet"

        wide.to_parquet(wide_path)

        long = (
            wide.stack(["ticker", "field"], future_stack=True)
            .rename("value")
            .reset_index()[["date", "ticker", "field", "value"]]
            .dropna(subset=["value"])
        )
        long.to_parquet(long_path, index=False)

        span = f"{wide.index.min().date()} -> {wide.index.max().date()}"
        print(f"  {group}: {wide.shape[1]} tickers, {len(wide)} rows, {span}")

    print("Done.")


if __name__ == "__main__":
    convert()
